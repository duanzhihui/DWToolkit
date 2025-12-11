#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
casegen - 任务清单转测试用例工具

从task.xlsx读取任务，根据YAML配置文件生成case.xlsx
"""

import argparse
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

import yaml
from openpyxl import Workbook, load_workbook


def load_config(config_path: str) -> Dict[str, Any]:
    """加载YAML配置文件"""
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def read_tasks(file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    读取任务清单Excel文件
    
    Args:
        file_path: Excel文件路径
        sheet_name: sheet名称，None则读取第一个sheet
        
    Returns:
        任务列表，每个任务是一个字典
    """
    wb = load_workbook(file_path, read_only=True)
    
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    
    # 第一行作为表头
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    
    tasks = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        task = {}
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else None
            task[header] = value if value is not None else ""
        tasks.append(task)
    
    wb.close()
    return tasks


def replace_placeholders(template: str, task: Dict[str, Any], context: Dict[str, Any]) -> str:
    """
    替换模板中的占位符
    
    Args:
        template: 模板字符串
        task: 任务数据
        context: 上下文变量（如index, task_index等）
        
    Returns:
        替换后的字符串
    """
    result = str(template)
    
    # 替换 <字段名> 格式的占位符（从任务中提取）
    pattern = r'<([^>]+)>'
    matches = re.findall(pattern, result)
    for field_name in matches:
        value = task.get(field_name, "")
        result = result.replace(f'<{field_name}>', str(value))
    
    # 替换 {变量名} 格式的占位符（上下文变量）
    for key, value in context.items():
        result = result.replace(f'{{{key}}}', str(value))
    
    return result


def generate_cases(tasks: List[Dict[str, Any]], config: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    根据配置生成测试用例
    
    Args:
        tasks: 任务列表
        config: 配置信息
        
    Returns:
        测试用例列表
    """
    cases_config = config.get('cases', [])
    if not cases_config:
        print("警告: 配置文件中没有定义用例模板")
        return []
    
    all_cases = []
    global_index = 1  # 全局用例序号
    
    for task_index, task in enumerate(tasks, start=1):
        for case_template in cases_config:
            case = {}
            context = {
                'index': global_index,
                'task_index': task_index,
            }
            
            fields = case_template.get('fields', [])
            for field in fields:
                field_name = field.get('name', '')
                field_value = field.get('value', '')
                case[field_name] = replace_placeholders(field_value, task, context)
            
            all_cases.append(case)
            global_index += 1
    
    return all_cases


def get_field_names(config: Dict[str, Any]) -> List[str]:
    """从配置中获取所有字段名称（保持顺序）"""
    cases_config = config.get('cases', [])
    if not cases_config:
        return []
    
    # 使用第一个用例模板的字段顺序
    first_template = cases_config[0]
    fields = first_template.get('fields', [])
    return [f.get('name', '') for f in fields]


def write_cases(cases: List[Dict[str, Any]], file_path: str, 
                sheet_name: str, field_names: List[str]) -> None:
    """
    将测试用例写入Excel文件
    
    Args:
        cases: 测试用例列表
        file_path: 输出文件路径
        sheet_name: sheet名称
        field_names: 字段名称列表（决定列顺序）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 写入表头
    for col, name in enumerate(field_names, start=1):
        ws.cell(row=1, column=col, value=name)
    
    # 写入数据
    for row_idx, case in enumerate(cases, start=2):
        for col_idx, field_name in enumerate(field_names, start=1):
            value = case.get(field_name, "")
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(file_path)
    print(f"已生成 {len(cases)} 条测试用例，保存至: {file_path}")


def main():
    parser = argparse.ArgumentParser(
        description='casegen - 任务清单转测试用例工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例:
  python casegen.py config.yml
  python casegen.py -c config.yml
        '''
    )
    parser.add_argument(
        'config', 
        nargs='?',
        help='YAML配置文件路径'
    )
    parser.add_argument(
        '-c', '--config-file',
        dest='config_file',
        help='YAML配置文件路径（可选方式）'
    )
    
    args = parser.parse_args()
    
    # 确定配置文件路径
    config_path = args.config or args.config_file
    if not config_path:
        parser.print_help()
        sys.exit(1)
    
    # 检查配置文件是否存在
    if not Path(config_path).exists():
        print(f"错误: 配置文件不存在: {config_path}")
        sys.exit(1)
    
    # 加载配置
    print(f"加载配置文件: {config_path}")
    config = load_config(config_path)
    
    # 获取配置项
    input_file = config.get('input_file', 'task.xlsx')
    output_file = config.get('output_file', 'case.xlsx')
    input_sheet = config.get('input_sheet')
    output_sheet = config.get('output_sheet', '测试用例')
    
    # 处理相对路径（相对于配置文件所在目录）
    config_dir = Path(config_path).parent
    if not Path(input_file).is_absolute():
        input_file = str(config_dir / input_file)
    if not Path(output_file).is_absolute():
        output_file = str(config_dir / output_file)
    
    # 检查输入文件是否存在
    if not Path(input_file).exists():
        print(f"错误: 任务清单文件不存在: {input_file}")
        sys.exit(1)
    
    # 读取任务
    print(f"读取任务清单: {input_file}")
    tasks = read_tasks(input_file, input_sheet)
    print(f"共读取 {len(tasks)} 条任务")
    
    if not tasks:
        print("警告: 任务清单为空")
        sys.exit(0)
    
    # 生成测试用例
    print("生成测试用例...")
    cases = generate_cases(tasks, config)
    
    if not cases:
        print("警告: 未生成任何测试用例")
        sys.exit(0)
    
    # 获取字段名称
    field_names = get_field_names(config)
    
    # 写入输出文件
    write_cases(cases, output_file, output_sheet, field_names)
    
    print("完成!")


if __name__ == '__main__':
    main()

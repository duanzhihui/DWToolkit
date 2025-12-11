#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""创建示例task.xlsx文件"""

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "任务清单"

# 表头
headers = ["任务编号", "任务名称", "所属模块", "前置条件", "测试步骤", "预期结果"]
for col, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col, value=header)

# 示例数据
tasks = [
    ["T001", "用户登录功能", "用户模块", "用户已注册", "1.输入用户名\n2.输入密码\n3.点击登录", "登录成功，跳转到首页"],
    ["T002", "用户注册功能", "用户模块", "无", "1.输入用户名\n2.输入密码\n3.确认密码\n4.点击注册", "注册成功，提示注册成功"],
    ["T003", "商品搜索功能", "商品模块", "商品数据已存在", "1.输入搜索关键词\n2.点击搜索", "显示匹配的商品列表"],
]

for row_idx, task in enumerate(tasks, start=2):
    for col_idx, value in enumerate(task, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

wb.save("task.xlsx")
print("已创建示例文件: task.xlsx")

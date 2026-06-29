#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
excel_merge - Excel 文件合并工具

根据 YAML 配置文件或命令行参数，合并多个 Excel 文件中指定 sheet 的数据。

特性:
  1. 输入输出通过配置文件或命令行参数指定
  2. 可配置需要合并的 sheet
  3. 可配置每个 sheet 数据起始行，支持多级表头
  4. 只合并数据，表头默认取第一个文件
  5. 可配置参与合并的 Excel 文件: 支持目录扫描 + 文件名正则匹配
"""

import argparse
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

import yaml
from openpyxl import Workbook, load_workbook


# Excel 临时文件前缀（Office 打开文件时生成的锁文件）
_TEMP_PREFIX = "~$"


def load_config(config_path: str) -> Dict[str, Any]:
    """加载 YAML 配置文件"""
    with open(config_path, 'r', encoding='utf-8') as f:
        try:
            return yaml.safe_load(f) or {}
        except yaml.YAMLError as e:
            print(f"错误: 配置文件 YAML 解析失败: {config_path}")
            print(f"  {e}")
            print("提示: Windows 路径请勿放在双引号中（反斜杠会被当作转义符）。")
            print("      可改用单引号 'D:\\dir'、正斜杠 \"D:/dir\" 或双反斜杠 \"D:\\\\dir\"。")
            sys.exit(1)


def collect_input_files(
    files: Optional[List[str]],
    directory: Optional[str],
    pattern: Optional[str],
    recursive: bool,
    base_dir: Path,
) -> List[Path]:
    """
    收集参与合并的 Excel 文件。

    Args:
        files: 显式指定的文件列表（相对 base_dir 或绝对路径）
        directory: 扫描目录
        pattern: 文件名正则表达式（匹配文件名，不含路径）
        recursive: 是否递归扫描子目录
        base_dir: 相对路径的基准目录

    Returns:
        去重并排序后的文件路径列表
    """
    result: List[Path] = []
    seen = set()

    def _add(p: Path) -> None:
        rp = p.resolve()
        if rp in seen:
            return
        if rp.name.startswith(_TEMP_PREFIX):
            return
        seen.add(rp)
        result.append(p)

    # 1. 显式文件列表
    for f in files or []:
        fp = Path(f)
        if not fp.is_absolute():
            fp = base_dir / fp
        if fp.exists():
            _add(fp)
        else:
            print(f"警告: 指定文件不存在，已跳过: {fp}")

    # 2. 目录扫描 + 正则匹配
    if directory:
        dir_path = Path(directory)
        if not dir_path.is_absolute():
            dir_path = base_dir / dir_path
        if not dir_path.is_dir():
            print(f"警告: 扫描目录不存在，已跳过: {dir_path}")
        else:
            regex = None
            if pattern:
                try:
                    regex = re.compile(pattern)
                except re.error as e:
                    print(f"错误: 文件名正则表达式无效: {pattern!r}")
                    print(f"  {e}")
                    print("提示: pattern 使用正则语法，非通配符。")
                    print(r'      匹配全部 xlsx: "\.xlsx$"；以"销售"开头: "^销售.*\.xlsx$"')
                    sys.exit(1)
            globber = dir_path.rglob('*') if recursive else dir_path.glob('*')
            candidates = sorted(
                (p for p in globber if p.is_file() and p.suffix.lower() in ('.xlsx', '.xlsm')),
                key=lambda p: str(p).lower(),
            )
            for p in candidates:
                if regex is None or regex.search(p.name):
                    _add(p)

    return result


def read_sheet_rows(
    file_path: Path,
    sheet_name: str,
    header_rows: int,
    data_start_row: Optional[int],
) -> tuple:
    """
    读取单个文件中指定 sheet 的表头和数据。

    Args:
        file_path: Excel 文件路径
        sheet_name: sheet 名称
        header_rows: 表头行数（支持多级表头）
        data_start_row: 数据起始行（1 基），None 则为 header_rows + 1

    Returns:
        (header_rows_data, data_rows) 元组；若 sheet 不存在返回 (None, None)
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return None, None
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if data_start_row is None:
        data_start_row = header_rows + 1

    header = rows[:header_rows] if header_rows > 0 else []
    data = rows[data_start_row - 1:]

    # 过滤完全空白的行
    data = [r for r in data if any(c is not None and str(c).strip() != "" for c in r)]
    return header, data


def merge_sheet(
    input_files: List[Path],
    sheet_cfg: Dict[str, Any],
) -> tuple:
    """
    合并所有文件中某个 sheet 的数据。

    Args:
        input_files: 输入文件列表
        sheet_cfg: 单个 sheet 的配置

    Returns:
        (output_sheet_name, header_rows_data, merged_data_rows)
    """
    name = sheet_cfg['name']
    output_name = sheet_cfg.get('output_name', name)
    header_rows = int(sheet_cfg.get('header_rows', 1))
    data_start_row = sheet_cfg.get('data_start_row')
    if data_start_row is not None:
        data_start_row = int(data_start_row)

    header_data = None
    merged_data: List[Any] = []
    used_files = 0

    for fp in input_files:
        header, data = read_sheet_rows(fp, name, header_rows, data_start_row)
        if header is None and data is None:
            print(f"  - {fp.name}: 未找到 sheet '{name}'，已跳过")
            continue
        used_files += 1
        # 表头取第一个含该 sheet 的文件
        if header_data is None:
            header_data = header
        merged_data.extend(data)
        print(f"  - {fp.name}: 合并 {len(data)} 行")

    if used_files == 0:
        print(f"  警告: 没有文件包含 sheet '{name}'")
    return output_name, header_data or [], merged_data


def write_output(
    output_path: Path,
    merged_sheets: List[tuple],
) -> None:
    """将合并结果写入输出 Excel 文件"""
    wb = Workbook()
    # 删除默认创建的空 sheet
    default_ws = wb.active
    wb.remove(default_ws)

    for output_name, header_data, data_rows in merged_sheets:
        ws = wb.create_sheet(title=output_name[:31])  # Excel sheet 名上限 31 字符
        row_idx = 1
        for header_row in header_data:
            for col_idx, value in enumerate(header_row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1
        for data_row in data_rows:
            for col_idx, value in enumerate(data_row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1

    if not wb.sheetnames:
        wb.create_sheet(title="Sheet1")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_config(args: argparse.Namespace) -> tuple:
    """合并配置文件与命令行参数，命令行优先。返回 (config, base_dir)"""
    config: Dict[str, Any] = {}
    base_dir = Path.cwd()

    config_path = args.config or args.config_file
    if config_path:
        if not Path(config_path).exists():
            print(f"错误: 配置文件不存在: {config_path}")
            sys.exit(1)
        print(f"加载配置文件: {config_path}")
        config = load_config(config_path)
        base_dir = Path(config_path).resolve().parent

    # 命令行参数覆盖
    if args.directory is not None:
        config['directory'] = args.directory
    if args.pattern is not None:
        config['pattern'] = args.pattern
    if args.files:
        config['files'] = args.files
    if args.output is not None:
        config['output'] = args.output
    if args.recursive:
        config['recursive'] = True
    if args.sheets:
        config['sheets'] = [{'name': s} for s in args.sheets]

    return config, base_dir


def main():
    parser = argparse.ArgumentParser(
        description='excel_merge - Excel 文件合并工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例:
  python excel_merge.py config.yml
  python excel_merge.py -c config.yml
  python excel_merge.py -d ./data -p "^销售.*\\.xlsx$" -s Sheet1 -o merged.xlsx
        '''
    )
    parser.add_argument('config', nargs='?', help='YAML 配置文件路径')
    parser.add_argument('-c', '--config-file', dest='config_file', help='YAML 配置文件路径（可选方式）')
    parser.add_argument('-d', '--directory', help='扫描目录')
    parser.add_argument('-p', '--pattern', help='文件名正则表达式')
    parser.add_argument('-f', '--files', nargs='*', help='显式指定的文件列表')
    parser.add_argument('-s', '--sheets', nargs='*', help='需要合并的 sheet 名称列表')
    parser.add_argument('-o', '--output', help='输出文件路径')
    parser.add_argument('-r', '--recursive', action='store_true', help='递归扫描子目录')

    args = parser.parse_args()

    if not (args.config or args.config_file or args.directory or args.files):
        parser.print_help()
        sys.exit(1)

    config, base_dir = build_config(args)

    # 解析配置项
    files = config.get('files')
    directory = config.get('directory')
    pattern = config.get('pattern')
    recursive = bool(config.get('recursive', False))
    output = config.get('output', 'merged.xlsx')
    sheets_cfg = config.get('sheets')

    if not sheets_cfg:
        print("错误: 未配置需要合并的 sheet（config.sheets 或 -s 参数）")
        sys.exit(1)

    # 收集输入文件
    print("收集输入文件...")
    input_files = collect_input_files(files, directory, pattern, recursive, base_dir)
    if not input_files:
        print("错误: 未找到任何输入文件")
        sys.exit(1)
    print(f"共找到 {len(input_files)} 个输入文件:")
    for fp in input_files:
        print(f"  * {fp}")

    # 逐个 sheet 合并
    merged_sheets = []
    for sheet_cfg in sheets_cfg:
        if isinstance(sheet_cfg, str):
            sheet_cfg = {'name': sheet_cfg}
        name = sheet_cfg.get('name')
        if not name:
            print("警告: 存在缺少 name 的 sheet 配置，已跳过")
            continue
        print(f"合并 sheet: {name}")
        merged_sheets.append(merge_sheet(input_files, sheet_cfg))

    # 输出
    output_path = Path(output)
    if not output_path.is_absolute():
        output_path = base_dir / output_path
    write_output(output_path, merged_sheets)

    total = sum(len(d) for _, _, d in merged_sheets)
    print(f"完成! 共合并 {total} 行数据，保存至: {output_path}")


if __name__ == '__main__':
    main()

"""pytest 公共夹具：用于生成临时 Excel 文件。"""

from pathlib import Path
from typing import Dict, List

import pytest
from openpyxl import Workbook


def write_xlsx(path: Path, sheets: Dict[str, List[list]]) -> Path:
    """按 ``{sheet名: [行, ...]}`` 生成 xlsx 文件。"""
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    if not wb.sheetnames:
        wb.create_sheet(title="Sheet1")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


@pytest.fixture
def make_xlsx(tmp_path):
    """返回一个在 tmp_path 下生成 xlsx 的工厂函数。"""

    def _factory(name: str, sheets: Dict[str, List[list]]) -> Path:
        return write_xlsx(tmp_path / name, sheets)

    return _factory

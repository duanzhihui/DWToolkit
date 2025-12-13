import os
import pandas as pd
from datetime import datetime
import argparse
import hashlib
import fnmatch
import yaml
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import xlwings as xw

class filetag:
    def __init__(self, folder_path, output_file=None, include_patterns=None, exclude_patterns=None):
        self.folder_path = folder_path
        self.include_patterns = include_patterns or []
        self.exclude_patterns = exclude_patterns or []
        # 处理输出文件路径
        if output_file:
            if os.path.isabs(output_file):
                self.excel_file = output_file
            else:
                self.excel_file = os.path.join(folder_path, output_file)
        else:
            self.excel_file = os.path.join(folder_path, "文件列表.xlsx")
        self.sheet_name = "文件列表"
        self.columns = ["文件目录", "文件名", "文件扩展名", "文件大小(MB)", 
                       "创建日期", "修改日期", "文件md5", "文件重复数",
                       "文件标签", "文件备注", "文件状态", "更新日期"]
        # 设置默认列宽
        self.default_widths = {
            "文件目录": 50,
            "文件名": 30,
            "文件扩展名": 12,
            "文件大小(MB)": 12,
            "创建日期": 20,
            "修改日期": 20,
            "文件md5": 35,
            "文件重复数": 12,
            "文件标签": 20,
            "文件备注": 30,
            "文件状态": 10,
            "更新日期": 12
        }

    def get_file_md5(self, file_path):
        """计算文件的md5值"""
        hasher = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hasher.update(chunk)
        return hasher.hexdigest()
    
    def get_file_size_mb(self, file_path):
        """获取文件大小（MB）"""
        size_bytes = os.path.getsize(file_path)
        return round(size_bytes / (1024 * 1024), 2)

    def get_file_dates(self, file_path):
        """获取文件的创建时间和修改时间"""
        # 获取文件状态信息
        stats = os.stat(file_path)
        # 转换时间戳为日期字符串
        create_time = datetime.fromtimestamp(stats.st_ctime).strftime('%Y-%m-%d %H:%M:%S')
        modify_time = datetime.fromtimestamp(stats.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
        return create_time, modify_time

    def get_file_extension(self, filename):
        """获取文件扩展名（不包含点号）"""
        _, ext = os.path.splitext(filename)
        return ext[1:] if ext else ""  # 移除点号

    def match_pattern(self, name, pattern):
        """检查名称是否匹配模式（支持Unix通配符）"""
        pattern = pattern.strip()
        if not pattern:
            return False
        # 目录模式以反斜杠结尾
        is_dir_pattern = pattern.endswith('\\') or pattern.endswith('/')
        if is_dir_pattern:
            pattern = pattern.rstrip('\\/') 
        return fnmatch.fnmatch(name, pattern)

    def should_include_file(self, file_path, filename):
        """判断文件是否应该被包含"""
        # 获取相对于扫描目录的路径
        rel_path = os.path.relpath(file_path, self.folder_path)
        path_parts = rel_path.split(os.sep)
        
        # 检查排除规则
        for pattern in self.exclude_patterns:
            pattern = pattern.strip()
            if not pattern:
                continue
            is_dir_pattern = pattern.endswith('\\') or pattern.endswith('/')
            if is_dir_pattern:
                # 目录模式：检查路径中的任意目录是否匹配
                dir_pattern = pattern.rstrip('\\/')
                for part in path_parts[:-1]:  # 排除文件名本身
                    if fnmatch.fnmatch(part, dir_pattern):
                        return False
            else:
                # 文件模式：检查文件名是否匹配
                if fnmatch.fnmatch(filename, pattern):
                    return False
        
        # 如果没有包含规则，默认包含所有
        if not self.include_patterns:
            return True
        
        # 检查包含规则
        for pattern in self.include_patterns:
            pattern = pattern.strip()
            if not pattern:
                continue
            is_dir_pattern = pattern.endswith('\\') or pattern.endswith('/')
            if is_dir_pattern:
                # 目录模式：检查路径中的任意目录是否匹配
                dir_pattern = pattern.rstrip('\\/')
                for part in path_parts[:-1]:
                    if fnmatch.fnmatch(part, dir_pattern):
                        return True
            else:
                # 文件模式：检查文件名是否匹配
                if fnmatch.fnmatch(filename, pattern):
                    return True
        
        return False

    def get_all_files(self):
        """递归获取文件夹下所有文件"""
        all_files = []
        for root, _, files in os.walk(self.folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                # 应用过滤规则
                if not self.should_include_file(file_path, file):
                    continue
                create_time, modify_time = self.get_file_dates(file_path)
                all_files.append({
                    "文件目录": os.path.dirname(file_path),
                    "文件名": file,
                    "文件扩展名": self.get_file_extension(file),
                    "文件大小(MB)": self.get_file_size_mb(file_path),
                    "创建日期": create_time,
                    "修改日期": modify_time,
                    "文件md5": self.get_file_md5(file_path),
                    "文件重复数": "",  # 将在Excel中用公式计算
                    "文件标签": "",
                    "文件备注": "",
                    "文件状态": "新增",
                    "更新日期": datetime.now().strftime("%Y%m%d")
                })
        return all_files

    def load_existing_data(self):
        """加载已存在的Excel文件数据"""
        if os.path.exists(self.excel_file):
            try:
                return pd.read_excel(self.excel_file, sheet_name=self.sheet_name)
            except Exception:
                return pd.DataFrame(columns=self.columns)
        return pd.DataFrame(columns=self.columns)

    def get_column_widths(self, worksheet):
        """获取现有Excel文件的列宽"""
        widths = {}
        for i, col_name in enumerate(self.columns, 1):
            widths[col_name] = worksheet.column_dimensions[get_column_letter(i)].width
        return widths

    def set_column_widths(self, worksheet, widths=None):
        """设置Excel列宽"""
        # 如果没有提供宽度，使用默认值
        if widths is None:
            widths = self.default_widths

        # 获取列索引
        header_indexes = {}
        for i, col_name in enumerate(self.columns, 1):
            if col_name in widths:
                header_indexes[col_name] = i

        # 设置列宽
        for header, width in widths.items():
            if header in header_indexes:
                worksheet.column_dimensions[get_column_letter(header_indexes[header])].width = width

    def add_hyperlinks_and_formulas(self, existing_widths=None):
        """为Excel文件添加超链接和公式"""
        wb = load_workbook(self.excel_file)
        ws = wb[self.sheet_name]
        
        # 获取列索引
        dir_col = None      # 文件目录列
        filename_col = None # 文件名列
        md5_col = None  # md5列
        duplicate_col = None # 重复数列
        tag_col = None     # 标签列
        
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "文件目录":
                dir_col = idx
            elif cell.value == "文件名":
                filename_col = idx
            elif cell.value == "文件md5":
                md5_col = idx
            elif cell.value == "文件重复数":
                duplicate_col = idx
            elif cell.value == "文件标签":
                tag_col = idx
        
        if not all([dir_col, filename_col, md5_col, duplicate_col, tag_col]):
            return
            
        # 获取数据范围
        last_row = ws.max_row
        md5_col_letter = get_column_letter(md5_col)
        
         
        # 添加文件重复数公式和超链接
        for row in range(2, last_row + 1):
            # COUNTIF公式计算重复数
            duplicate_cell = ws.cell(row=row, column=duplicate_col)
            duplicate_cell.value = f'=COUNTIF({md5_col_letter}2:{md5_col_letter}{last_row},{md5_col_letter}{row})'
            
            # 获取文件路径组件
            dir_value = ws.cell(row=row, column=dir_col).value
            filename_value = ws.cell(row=row, column=filename_col).value
            
            # 只有当两个值都不为空时才构建文件路径
            if dir_value and filename_value:
                # 添加文件超链接
                file_path = os.path.join(dir_value, filename_value)
                
                if os.path.exists(file_path):
                    cell = ws.cell(row=row, column=filename_col)
                    cell.hyperlink = file_path
                    cell.style = 'Hyperlink'
        
        
        # 设置列宽
        self.set_column_widths(ws, existing_widths)
        
        # 冻结面板
        ws.freeze_panes = 'D2'
        
        wb.save(self.excel_file)

    def create_table_and_slicer(self):
        """创建表格和标签切片器"""
        try:
            # 打开Excel应用程序
            app = xw.App(visible=False)
            wb = app.books.open(os.path.abspath(self.excel_file))
            
            try:
                # 获取主工作表
                ws = wb.sheets[self.sheet_name]
                
                # 获取数据范围
                data_range = ws.range("A1").end('down').end('right')
                
                # 尝试获取现有表格，如果不存在则创建新表格
                try:
                    table = ws.api.ListObjects.Item("FileTable")
                except:
                    # 创建新表格
                    start_cell = ws.range("A1")
                    end_cell = ws.range(data_range.row, len(self.columns))
                    table_range = ws.range(start_cell, end_cell)
                    
                    table = ws.api.ListObjects.Add(
                        SourceType=1,  # xlSrcRange
                        Source=table_range.api,
                        XlListObjectHasHeaders=1,  # xlYes
                        Destination=table_range.api
                    )
                    table.Name = "FileTable"
                    table.TableStyle = "TableStyleMedium2"
                
                # 获取标签列
                tag_col_index = None
                for i in range(1, table.ListColumns.Count + 1):
                    col = table.ListColumns.Item(i)
                    if col.Name == "文件标签":
                        tag_col_index = i
                        break
                
                if tag_col_index:
                    try:
                        # 尝试获取现有的切片器缓存
                        slicer_cache = None
                        for cache in wb.api.SlicerCaches:
                            if cache.SourceName == "FileTable" and cache.SlicerFields(1).Name == "文件标签":
                                slicer_cache = cache
                                break
                        
                        if not slicer_cache:
                            # 创建新的切片器缓存
                            slicer_cache = wb.api.SlicerCaches.Add(
                                table,
                                "文件标签"
                            )
                            # 添加切片器到工作表
                            slicer = slicer_cache.Slicers.Add(
                                ws.api,
                                Top=0,
                                Left=data_range.api.Width + 20,
                                Width=120,
                                Height=200
                            )
                            slicer.Style = "SlicerStyleLight1"
                    
                    except Exception as e:
                        print(f"添加切片器时出错: {str(e)}")
                
                # 保存表格
                wb.save()
                
            finally:
                # 关闭工作簿和应用程序
                wb.close()
                app.quit()
                
        except Exception as e:
            print(f"创建表格和切片器时出错: {str(e)}")
            try:
                wb.close()
                app.quit()
            except:
                pass

    def update_file_list(self):
        """更新文件列表"""
        # 获取现有Excel文件的列宽
        existing_widths = None
        if os.path.exists(self.excel_file):
            wb = load_workbook(self.excel_file)
            ws = wb[self.sheet_name]
            existing_widths = self.get_column_widths(ws)
            wb.close()
        
        # 获取当前所有文件
        current_files = self.get_all_files()
        current_df = pd.DataFrame(current_files)
        
        # 加载已存在的数据
        existing_df = self.load_existing_data()
        
        if not existing_df.empty:
            # 构建文件唯一标识（目录+文件名）
            current_df['file_id'] = current_df['文件目录'] + '/' + current_df['文件名']
            existing_df['file_id'] = existing_df['文件目录'] + '/' + existing_df['文件名']
            
            # 标记已存在的文件
            for idx, row in current_df.iterrows():
                if row['file_id'] in existing_df['file_id'].values:
                    # 保持原有的标签和备注
                    existing_row = existing_df[existing_df['file_id'] == row['file_id']].iloc[0]
                    current_df.at[idx, '文件标签'] = existing_row['文件标签']
                    current_df.at[idx, '文件备注'] = existing_row['文件备注']
                    current_df.at[idx, '文件状态'] = "已有"
                    current_df.at[idx, '更新日期'] = existing_row['更新日期']
            
            # 标记删除的文件
            deleted_files = existing_df[~existing_df['file_id'].isin(current_df['file_id'])].copy()
            if not deleted_files.empty:
                deleted_files['文件状态'] = "删除"
                deleted_files['更新日期'] = datetime.now().strftime("%Y%m%d")
                current_df = pd.concat([current_df, deleted_files])
            
            # 删除临时列
            current_df = current_df.drop('file_id', axis=1)
        
        # 保存到Excel
        current_df.to_excel(self.excel_file, sheet_name=self.sheet_name, index=False)
        
        # 添加超链接和公式
        self.add_hyperlinks_and_formulas(existing_widths)
        
        # 创建表格和切片器
        self.create_table_and_slicer()
def load_config(config_path):
    """加载配置文件"""
    default_config = {
        'directory': '',
        'output': '文件列表.xlsx',
        'include': '',
        'exclude': ''
    }
    
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = yaml.safe_load(f) or {}
            # 合并默认配置
            for key in default_config:
                if key not in config:
                    config[key] = default_config[key]
            return config
        except Exception as e:
            print(f"读取配置文件失败: {e}")
    return default_config

def parse_patterns(pattern_str):
    """解析过滤模式字符串，以分号分隔"""
    if not pattern_str:
        return []
    return [p.strip() for p in pattern_str.split(';') if p.strip()]

def main():
    # 获取程序所在目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, 'config.yml')
    
    # 加载配置文件
    config = load_config(config_path)
    
    # 创建命令行参数解析器
    parser = argparse.ArgumentParser(description='文件标签管理')
    parser.add_argument('--dir', '-d', 
                      default=None,
                      help='要扫描的目录路径，默认为配置文件中的目录或程序所在目录')
    parser.add_argument('--include', '-i',
                      default=None,
                      help='包含过滤器，支持Unix通配符，多个规则用分号分隔')
    parser.add_argument('--exclude', '-e',
                      default=None,
                      help='排除过滤器，支持Unix通配符，多个规则用分号分隔')
    parser.add_argument('--output', '-o',
                      default=None,
                      help='输出文件路径，默认为扫描目录下的"文件列表.xlsx"')
    parser.add_argument('--config', '-c',
                      default=None,
                      help='指定配置文件路径')
    args = parser.parse_args()
    
    # 如果指定了配置文件，重新加载
    if args.config:
        config = load_config(args.config)
    
    # 确定扫描目录（命令行 > 配置文件 > 程序目录）
    if args.dir:
        scan_dir = args.dir
    elif config.get('directory'):
        scan_dir = config['directory']
    else:
        scan_dir = script_dir
    
    # 确定过滤规则（命令行 > 配置文件）
    include_str = args.include if args.include is not None else config.get('include', '')
    exclude_str = args.exclude if args.exclude is not None else config.get('exclude', '')
    
    include_patterns = parse_patterns(include_str)
    exclude_patterns = parse_patterns(exclude_str)
    
    # 确定输出文件路径（命令行 > 配置文件 > 默认值）
    output_file = args.output if args.output is not None else config.get('output', '文件列表.xlsx')
    
    # 打印配置信息
    print(f"扫描目录: {scan_dir}")
    if include_patterns:
        print(f"包含规则: {'; '.join(include_patterns)}")
    if exclude_patterns:
        print(f"排除规则: {'; '.join(exclude_patterns)}")
    
    # 使用传入的目录路径和过滤规则
    manager = filetag(scan_dir, output_file, include_patterns, exclude_patterns)
    manager.update_file_list()
    print(f"文件列表已更新，保存在：{manager.excel_file}")

if __name__ == "__main__":
    main()
